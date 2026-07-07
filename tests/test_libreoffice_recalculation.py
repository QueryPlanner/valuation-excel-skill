from __future__ import annotations

import io
import json
import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import recalculate_with_libreoffice
from fill_excel import fill_valuation_excel
from recalculate_with_libreoffice import (
    LibreOfficeRecalculationError,
    _prepare_local_input,
    _resolve_executable,
    _set_xml_attribute,
    recalculate_and_verify,
    restore_calculation_flags,
)
from run_valuation import run_valuation
from validate_inputs import normalize_and_validate_inputs
from workbook_contract import load_workbook_contract

from tests.helpers import TEMPLATE, VALID_INPUTS, load_valid_inputs, write_json


class LibreOfficeRecalculationTests(unittest.TestCase):
    def _filled_workbook(self, directory: Path) -> tuple[Path, Path]:
        normalized = normalize_and_validate_inputs(
            load_valid_inputs(),
            load_workbook_contract(TEMPLATE),
        )
        normalized_path = directory / "inputs.normalized.json"
        write_json(normalized_path, normalized)
        workbook_path = directory / "valuation.awaiting-recalculation.xlsx"
        fill_valuation_excel(
            "Example Foods",
            normalized_path,
            TEMPLATE,
            workbook_path,
            run_id="libreoffice-test",
        )
        return normalized_path, workbook_path

    def test_resolves_executable_and_updates_xml_attribute(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            executable = Path(temporary) / "soffice"
            executable.write_text("", encoding="utf-8")
            self.assertEqual(_resolve_executable(executable), str(executable.resolve()))
        with patch("recalculate_with_libreoffice.shutil.which", return_value="/bin/soffice"):
            self.assertEqual(_resolve_executable("soffice"), "/bin/soffice")
            self.assertEqual(_resolve_executable(), "/bin/soffice")
        with patch("recalculate_with_libreoffice.shutil.which", return_value=None):
            with self.assertRaisesRegex(LibreOfficeRecalculationError, "not found"):
                _resolve_executable("missing")
            with self.assertRaisesRegex(LibreOfficeRecalculationError, "not installed"):
                _resolve_executable()
        self.assertEqual(
            _set_xml_attribute('<calcPr calcMode="manual"/>', "calcMode", "auto"),
            '<calcPr calcMode="auto"/>',
        )
        self.assertEqual(
            _set_xml_attribute("<calcPr/>", "forceFullCalc", "1"),
            '<calcPr forceFullCalc="1"/>',
        )

    def test_restores_calculation_flags_without_changing_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "book.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active["A1"] = "=1+1"
            workbook.calculation.calcMode = "manual"
            workbook.calculation.fullCalcOnLoad = False
            workbook.calculation.forceFullCalc = False
            workbook.calculation.iterate = False
            workbook.save(path)
            workbook.close()

            restore_calculation_flags(path)
            restored = openpyxl.load_workbook(path, data_only=False)
            self.assertEqual(restored.active["A1"].value, "=1+1")
            self.assertEqual(restored.calculation.calcMode, "auto")
            self.assertTrue(restored.calculation.fullCalcOnLoad)
            self.assertTrue(restored.calculation.forceFullCalc)
            self.assertTrue(restored.calculation.iterate)
            self.assertEqual(restored.calculation.iterateCount, 100)
            restored.close()

    def test_restore_rejects_missing_workbook_or_calculation_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            missing_workbook = directory / "missing-workbook.xlsx"
            with ZipFile(missing_workbook, "w", compression=ZIP_DEFLATED) as archive:
                archive.writestr("placeholder.txt", "content")
            with self.assertRaisesRegex(
                LibreOfficeRecalculationError,
                "does not contain xl/workbook.xml",
            ):
                restore_calculation_flags(missing_workbook)

            missing_calculation = directory / "missing-calculation.xlsx"
            with ZipFile(missing_calculation, "w", compression=ZIP_DEFLATED) as archive:
                archive.writestr("xl/workbook.xml", "<workbook/>")
            with self.assertRaisesRegex(
                LibreOfficeRecalculationError,
                "missing xl/workbook.xml calcPr",
            ):
                restore_calculation_flags(missing_calculation)

    def test_template_guards_dormant_libreoffice_formulas(self) -> None:
        workbook = openpyxl.load_workbook(TEMPLATE, data_only=False)
        self.assertEqual(
            workbook["Cost of capital worksheet"]["J21"].value,
            "=IF($H$32=0,0,H21/$H$32)",
        )
        self.assertTrue(
            workbook["Cost of capital worksheet"]["K48"].value.startswith("=IF(J48=0,0,"),
        )
        self.assertTrue(
            workbook["Synthetic rating"]["D13"].value.startswith('=IF(C7=0,"N/A",'),
        )
        self.assertTrue(
            workbook["Synthetic rating"]["D14"].value.startswith("=IF(C7=0,0,"),
        )
        workbook.close()

    def test_prepares_complete_local_input_and_rejects_missing_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            _, source = self._filled_workbook(directory)
            prepared = directory / "prepared.xlsx"
            _prepare_local_input(source, prepared)
            workbook = openpyxl.load_workbook(prepared, data_only=False)
            self.assertEqual(workbook["Run Metadata"]["B2"].value, "complete")
            self.assertTrue(workbook.calculation.iterate)
            workbook.close()

            missing = directory / "missing.xlsx"
            workbook = openpyxl.Workbook()
            workbook.save(missing)
            workbook.close()
            with self.assertRaisesRegex(LibreOfficeRecalculationError, "Run Metadata"):
                _prepare_local_input(missing, directory / "unused.xlsx")

    def test_mocked_recalculation_writes_verified_receipt(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized, workbook = self._filled_workbook(directory)
            output = directory / "output"

            def fake_run(command, **_kwargs):
                converted_directory = Path(command[command.index("--outdir") + 1])
                shutil.copy2(command[-1], converted_directory / Path(command[-1]).name)
                return subprocess.CompletedProcess(command, 0, "converted", "")

            verification = {
                "status": "passed",
                "outputs": {
                    "wacc": 0.08,
                    "operating_assets": 1500.0,
                    "common_equity": 1200.0,
                    "value_per_share": 12.0,
                },
            }
            with (
                patch(
                    "recalculate_with_libreoffice._resolve_executable",
                    return_value="/bin/soffice",
                ),
                patch(
                    "recalculate_with_libreoffice.subprocess.run",
                    side_effect=fake_run,
                ),
                patch(
                    "recalculate_with_libreoffice.verify_recalculated",
                    return_value=verification,
                ),
            ):
                receipt = recalculate_and_verify(
                    excel_file_path=workbook,
                    normalized_inputs_path=normalized,
                    template_path=TEMPLATE,
                    output_directory=output,
                )
            self.assertEqual(receipt["status"], "complete")
            self.assertEqual(receipt["outputs"]["value_per_share"], 12.0)
            self.assertTrue(Path(receipt["snapshot_path"]).is_file())
            self.assertEqual(
                json.loads((output / "libreoffice-recalculation.json").read_text())["backend"],
                "libreoffice",
            )

            with self.assertRaises(FileExistsError):
                recalculate_and_verify(
                    excel_file_path=workbook,
                    normalized_inputs_path=normalized,
                    template_path=TEMPLATE,
                    output_directory=output,
                    executable="/bin/soffice",
                )

    def test_recalculation_surfaces_validation_and_process_failures(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized, workbook = self._filled_workbook(directory)
            with self.assertRaisesRegex(LibreOfficeRecalculationError, "greater than zero"):
                recalculate_and_verify(
                    excel_file_path=workbook,
                    normalized_inputs_path=normalized,
                    template_path=TEMPLATE,
                    output_directory=directory / "bad-timeout",
                    timeout_seconds=0,
                )
            with self.assertRaisesRegex(LibreOfficeRecalculationError, "does not exist"):
                recalculate_and_verify(
                    excel_file_path=directory / "missing.xlsx",
                    normalized_inputs_path=normalized,
                    template_path=TEMPLATE,
                    output_directory=directory / "missing",
                )

            common_patches = (
                patch(
                    "recalculate_with_libreoffice._resolve_executable",
                    return_value="/bin/soffice",
                ),
            )
            with (
                common_patches[0],
                patch(
                    "recalculate_with_libreoffice.subprocess.run",
                    side_effect=subprocess.TimeoutExpired(["soffice"], 1),
                ),
            ):
                with self.assertRaisesRegex(LibreOfficeRecalculationError, "exceeded"):
                    recalculate_and_verify(
                        excel_file_path=workbook,
                        normalized_inputs_path=normalized,
                        template_path=TEMPLATE,
                        output_directory=directory / "timeout",
                        timeout_seconds=1,
                    )

            with (
                patch(
                    "recalculate_with_libreoffice._resolve_executable",
                    return_value="/bin/soffice",
                ),
                patch(
                    "recalculate_with_libreoffice.subprocess.run",
                    return_value=subprocess.CompletedProcess(
                        ["soffice"],
                        2,
                        "",
                        "conversion failed",
                    ),
                ),
            ):
                with self.assertRaisesRegex(LibreOfficeRecalculationError, "exit code 2"):
                    recalculate_and_verify(
                        excel_file_path=workbook,
                        normalized_inputs_path=normalized,
                        template_path=TEMPLATE,
                        output_directory=directory / "failed",
                    )

            with (
                patch(
                    "recalculate_with_libreoffice._resolve_executable",
                    return_value="/bin/soffice",
                ),
                patch(
                    "recalculate_with_libreoffice.subprocess.run",
                    return_value=subprocess.CompletedProcess(["soffice"], 0, "done", ""),
                ),
            ):
                with self.assertRaisesRegex(LibreOfficeRecalculationError, "did not create"):
                    recalculate_and_verify(
                        excel_file_path=workbook,
                        normalized_inputs_path=normalized,
                        template_path=TEMPLATE,
                        output_directory=directory / "no-output",
                    )

    def test_main_success_and_failure(self) -> None:
        arguments = [
            "recalculate_with_libreoffice.py",
            "--file",
            "value.xlsx",
            "--inputs",
            "inputs.json",
            "--template",
            "template.xlsx",
            "--output-dir",
            "/tmp/output",
        ]
        with (
            patch(
                "recalculate_with_libreoffice.recalculate_and_verify",
                return_value={"status": "complete"},
            ),
            patch("sys.argv", arguments),
            patch("sys.stdout", new_callable=io.StringIO),
        ):
            recalculate_with_libreoffice.main()

        with (
            patch(
                "recalculate_with_libreoffice.recalculate_and_verify",
                side_effect=LibreOfficeRecalculationError("failed"),
            ),
            patch("sys.argv", arguments),
            patch("sys.stderr", new_callable=io.StringIO),
            self.assertRaises(SystemExit) as context,
        ):
            recalculate_with_libreoffice.main()
        self.assertEqual(context.exception.code, 1)

    @unittest.skipUnless(shutil.which("soffice"), "LibreOffice is not installed")
    def test_real_libreoffice_backend_recalculates_and_verifies(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            manifest, exit_code = run_valuation(
                inputs_path=VALID_INPUTS,
                template_path=TEMPLATE,
                output_root=Path(temporary) / "runs",
                backend="libreoffice",
                run_id="real-libreoffice",
            )
            self.assertEqual(exit_code, 0, manifest["errors"])
            self.assertEqual(manifest["status"], "complete")
            outputs = manifest["steps"]["recalculation_delivery"]["receipt"]["outputs"]
            self.assertGreater(outputs["value_per_share"], 0)
            self.assertGreater(outputs["wacc"], 0)


if __name__ == "__main__":
    unittest.main()
