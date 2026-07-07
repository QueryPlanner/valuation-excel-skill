from __future__ import annotations

import contextlib
import io
import json
import runpy
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from common import atomic_write_json
import fill_excel
import verify_workbook
from fill_excel import (
    _load_inputs as load_fill_inputs,
    _replace_sheet,
    _write_business_revenue,
    _write_geographic_revenue,
    _write_r_and_d_sheet,
    fill_valuation_excel,
    main as fill_main,
)
from validate_inputs import normalize_and_validate_inputs
from verify_workbook import (
    WorkbookVerificationError,
    _active_sheets,
    _load_inputs as load_verify_inputs,
    main as verify_main,
    verify_precalculation,
    verify_recalculated,
)
from workbook_contract import load_workbook_contract

from tests.helpers import (
    TEMPLATE,
    analyst_evidence,
    filing_evidence,
    inject_cached_values,
    load_valid_inputs,
    replace_evidence,
)


class FillAndVerifyTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.contract = load_workbook_contract(TEMPLATE)

    def _normalize(self, directory: Path, inputs: dict) -> Path:
        normalized = normalize_and_validate_inputs(inputs, self.contract)
        path = directory / "normalized.json"
        atomic_write_json(path, normalized)
        return path

    def _fill(self, directory: Path, inputs: dict, name: str = "valuation.xlsx") -> tuple[Path, Path]:
        normalized_path = self._normalize(directory, inputs)
        workbook_path = directory / name
        fill_valuation_excel(
            inputs["company_context"]["company_name"],
            normalized_path,
            TEMPLATE,
            workbook_path,
            run_id="test-run",
        )
        return normalized_path, workbook_path

    def test_basic_fill_writes_audit_metadata_and_calculation_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path, workbook_path = self._fill(directory, load_valid_inputs())
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)
            self.assertIn("Sources & Audit", workbook.sheetnames)
            self.assertIn("Run Metadata", workbook.sheetnames)
            self.assertEqual(workbook["Input sheet"]["B4"].value, "Example Foods")
            self.assertEqual(workbook["Cost of capital worksheet"]["B25"].value, "Will Input")
            self.assertEqual(workbook["Cost of capital worksheet"]["B26"].value, 0.05)
            self.assertEqual(workbook["Input sheet"]["B71"].value, 0)
            self.assertTrue(workbook.calculation.iterate)
            self.assertEqual(workbook.calculation.iterateCount, 100)
            workbook.close()
            report = verify_precalculation(workbook_path, normalized_path, TEMPLATE)
            self.assertEqual(report["status"], "passed")

    def test_refuses_local_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path, workbook_path = self._fill(directory, load_valid_inputs())
            with self.assertRaises(FileExistsError):
                fill_valuation_excel(
                    "Example Foods",
                    normalized_path,
                    TEMPLATE,
                    workbook_path,
                    run_id="second-run",
                )
            fill_valuation_excel(
                "Example Foods",
                normalized_path,
                TEMPLATE,
                workbook_path,
                run_id="second-run",
                replace=True,
            )

    def test_populates_optional_model_paths(self) -> None:
        inputs = load_valid_inputs()
        inputs["revenue_splits"].update(
            {
                "geography_mode": "country",
                "by_country": {"United States": 900.0, "Rest of the World": 100.0},
                "by_region": {},
                "by_business": {"Food Processing": 600.0, "Restaurant/Dining": 400.0},
            }
        )
        replace_evidence(
            inputs,
            "revenue_splits.by_business.Food Processing",
            filing_evidence("revenue_splits.by_business.Food Processing", 600.0),
        )
        inputs["source_evidence"].extend(
            (
                filing_evidence("revenue_splits.by_business.Restaurant/Dining", 400.0),
                filing_evidence("revenue_splits.by_country.United States", 900.0),
                filing_evidence("revenue_splits.by_country.Rest of the World", 100.0),
                analyst_evidence("cost_of_capital_inputs.rest_of_world_erp", 0.07),
            )
        )
        inputs["cost_of_capital_inputs"].update(
            {"debt_rating": "bbb", "direct_erp": None, "rest_of_world_erp": 0.07}
        )
        inputs["source_evidence"] = [
            record
            for record in inputs["source_evidence"]
            if record["metric"] != "cost_of_capital_inputs.direct_erp"
        ]
        inputs["r_and_d_details"] = {
            "current_year_expense": 30.0,
            "amortization_period_years": 3,
            "historical_expenses": {
                "Year_Minus_1": 25.0,
                "Year_Minus_2": 20.0,
                "Year_Minus_3": 15.0,
            },
        }
        for metric, value in (
            ("r_and_d_details.current_year_expense", 30.0),
            ("r_and_d_details.historical_expenses.Year_Minus_1", 25.0),
            ("r_and_d_details.historical_expenses.Year_Minus_2", 20.0),
            ("r_and_d_details.historical_expenses.Year_Minus_3", 15.0),
        ):
            inputs["source_evidence"].append(filing_evidence(metric, value))
        inputs["employee_options"] = {
            "total_options_outstanding": 5.0,
            "weighted_average_exercise_price": 10.0,
            "average_maturity_years": 4.0,
            "stock_price_standard_deviation": 0.3,
        }
        for field, value in inputs["employee_options"].items():
            inputs["source_evidence"].append(
                filing_evidence(f"employee_options.{field}", value)
            )
        inputs["operating_lease_details"] = {
            "capitalize": True,
            "book_debt_excludes_operating_leases": True,
            "current_year_expense": 12.0,
            "commitments": {
                "year_1": 10.0,
                "year_2": 9.0,
                "year_3": 8.0,
                "year_4": 7.0,
                "year_5": 6.0,
                "years_6_and_beyond": 20.0,
            },
        }
        inputs["source_evidence"].append(
            filing_evidence("operating_lease_details.current_year_expense", 12.0)
        )
        for key, value in inputs["operating_lease_details"]["commitments"].items():
            inputs["source_evidence"].append(
                filing_evidence(f"operating_lease_details.commitments.{key}", value)
            )
        inputs["advanced_assumptions"] = {
            "terminal_cost_of_capital_override": 0.08,
            "stable_return_on_capital_override": 0.1,
            "probability_of_failure": 0.05,
            "failure_proceeds_basis": "B",
            "failure_proceeds_percent": 0.4,
            "reinvestment_lag_years": 2,
            "hold_effective_tax_rate": True,
            "net_operating_loss": 10.0,
            "stable_riskfree_rate_override": 0.035,
            "terminal_growth_override": 0.03,
            "trapped_cash": 20.0,
            "trapped_cash_tax_rate": 0.1,
        }
        with tempfile.TemporaryDirectory() as temporary:
            _, workbook_path = self._fill(Path(temporary), inputs)
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)
            input_sheet = workbook["Input sheet"]
            self.assertEqual(input_sheet["B16"].value, "Yes")
            self.assertEqual(input_sheet["B17"].value, "Yes")
            self.assertEqual(input_sheet["B37"].value, "Yes")
            self.assertEqual(input_sheet["B45"].value, "Yes")
            self.assertEqual(input_sheet["B70"].value, "Yes")
            cost = workbook["Cost of capital worksheet"]
            self.assertEqual(cost["B25"].value, "Operating countries")
            self.assertEqual(cost["I16"].value, 0.07)
            self.assertEqual(cost["B21"].value, "Multibusiness(Global)")
            self.assertEqual(cost["B33"].value, "Actual rating")
            self.assertEqual(workbook["R& D converter"]["F7"].value, 30.0)
            self.assertEqual(workbook["Operating lease converter"]["B12"].value, 20.0)
            workbook.close()

    def test_recalculation_verification_rejects_blank_and_accepts_numeric_cache(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path, workbook_path = self._fill(directory, load_valid_inputs())
            with self.assertRaises(WorkbookVerificationError) as context:
                verify_recalculated(workbook_path, normalized_path, TEMPLATE)
            self.assertIn("wacc", "\n".join(context.exception.errors))

            recalculated = directory / "recalculated.xlsx"
            inject_cached_values(
                workbook_path,
                recalculated,
                {
                    "Cost of capital worksheet": {"B13": (0.08, None)},
                    "Valuation output": {
                        "B24": (1500.0, None),
                        "B31": (1200.0, None),
                        "B33": (12.0, None),
                    },
                },
            )
            report = verify_recalculated(recalculated, normalized_path, TEMPLATE)
            self.assertEqual(report["outputs"]["value_per_share"], 12.0)

    def test_verifier_detects_structure_formula_and_active_errors(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path, workbook_path = self._fill(directory, load_valid_inputs())
            broken = directory / "broken.xlsx"
            workbook = openpyxl.load_workbook(workbook_path)
            del workbook["Sources & Audit"]
            workbook["Input sheet"]["B7"] = "Wrong"
            workbook["Valuation output"]["C2"] = "=#REF!"
            workbook.calculation.iterate = False
            workbook.save(broken)
            workbook.close()
            with self.assertRaises(WorkbookVerificationError) as context:
                verify_precalculation(broken, normalized_path, TEMPLATE)
            errors = "\n".join(context.exception.errors)
            self.assertIn("audit sheets", errors)
            self.assertIn("fingerprint", errors)
            self.assertIn("iterative", errors)
            self.assertIn("expected", errors)
            self.assertIn("Formula text", errors)

            error_snapshot = directory / "error.xlsx"
            inject_cached_values(
                workbook_path,
                error_snapshot,
                {
                    "Cost of capital worksheet": {"B13": (0.08, None)},
                    "Valuation output": {
                        "B24": ("#VALUE!", "e"),
                        "B31": (1200.0, None),
                        "B33": (12.0, None),
                    },
                },
            )
            with self.assertRaises(WorkbookVerificationError) as active_context:
                verify_recalculated(error_snapshot, normalized_path, TEMPLATE)
            self.assertIn("formula errors", "\n".join(active_context.exception.errors))

    def test_fill_helpers_cover_geography_replacement_and_defensive_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            list_path = directory / "list.json"
            list_path.write_text("[]", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "must contain an object"):
                load_fill_inputs(list_path)

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            country_inputs = {
                "revenue_splits": {
                    "geography_mode": "country",
                    "by_country": {"United States": 100.0},
                    "by_region": {},
                    "by_business": {"Food Processing": 100.0},
                },
                "cost_of_capital_inputs": {"rest_of_world_erp": None},
            }
            self.assertEqual(
                _write_geographic_revenue(worksheet, country_inputs, {"North America": 21}),
                "Operating countries",
            )
            self.assertIsNone(worksheet["G16"].value)

            region_inputs = {
                "revenue_splits": {
                    "geography_mode": "region",
                    "by_country": {},
                    "by_region": {"North America": 100.0},
                    "by_business": {"Food Processing": 100.0},
                },
                "cost_of_capital_inputs": {"rest_of_world_erp": None},
            }
            self.assertEqual(
                _write_geographic_revenue(worksheet, region_inputs, {"North America": 21}),
                "Operating regions",
            )
            self.assertEqual(worksheet["H21"].value, 100.0)
            _write_business_revenue(worksheet, region_inputs)
            self.assertEqual(worksheet["B21"].value, "Single Business(Global)")

            workbook.create_sheet("replace-me")
            replacement = _replace_sheet(workbook, "replace-me")
            self.assertEqual(replacement.title, "replace-me")
            self.assertEqual(workbook.sheetnames.count("replace-me"), 1)
            workbook.close()

            template = openpyxl.load_workbook(TEMPLATE)
            del template["R& D converter"]
            with self.assertRaisesRegex(ValueError, "missing the R&D converter"):
                _write_r_and_d_sheet(template, load_valid_inputs())
            template.close()

    def test_fill_rejects_formula_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path = self._normalize(directory, load_valid_inputs())
            with patch.object(
                fill_excel,
                "protected_formula_fingerprint",
                side_effect=("before", "after"),
            ):
                with self.assertRaisesRegex(ValueError, "Protected workbook formulas changed"):
                    fill_valuation_excel(
                        "Example Foods",
                        normalized_path,
                        TEMPLATE,
                        directory / "not-written.xlsx",
                        run_id="mutation-test",
                    )

    def test_verifier_helpers_and_optional_active_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "list.json"
            path.write_text("[]", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "must contain an object"):
                load_verify_inputs(path)

        inputs = load_valid_inputs()
        inputs["r_and_d_details"]["current_year_expense"] = 1.0
        inputs["operating_lease_details"]["capitalize"] = True
        inputs["employee_options"]["total_options_outstanding"] = 1.0
        active = _active_sheets(inputs)
        self.assertIn("R& D converter", active)
        self.assertIn("Operating lease converter", active)
        self.assertIn("Option value", active)

    def test_verifier_detects_numeric_mismatch_flags_and_wacc_range(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path, workbook_path = self._fill(directory, load_valid_inputs())
            broken = directory / "bad-flags.xlsx"
            workbook = openpyxl.load_workbook(workbook_path)
            workbook["Input sheet"]["B22"] = 99.0
            workbook.calculation.calcMode = "manual"
            workbook.save(broken)
            workbook.close()
            with self.assertRaises(WorkbookVerificationError) as context:
                verify_precalculation(broken, normalized_path, TEMPLATE)
            errors = "\n".join(context.exception.errors)
            self.assertIn("full-recalculation flags", errors)
            self.assertIn("B22 expected", errors)

            bad_wacc = directory / "bad-wacc.xlsx"
            inject_cached_values(
                workbook_path,
                bad_wacc,
                {
                    "Cost of capital worksheet": {"B13": (0.75, None)},
                    "Valuation output": {
                        "B24": (1500.0, None),
                        "B31": (1200.0, None),
                        "B33": (12.0, None),
                    },
                },
            )
            with self.assertRaises(WorkbookVerificationError) as wacc_context:
                verify_recalculated(bad_wacc, normalized_path, TEMPLATE)
            self.assertIn("WACC must", "\n".join(wacc_context.exception.errors))

    def test_verify_main_success_failure_and_output_routing(self) -> None:
        report = {"stage": "precalculation", "status": "passed", "errors": []}
        with tempfile.TemporaryDirectory() as temporary:
            receipt = Path(temporary) / "report.json"
            argv = [
                "verify_workbook.py",
                "--file",
                "model.xlsx",
                "--inputs",
                "inputs.json",
                "--template",
                "template.xlsx",
                "--stage",
                "precalculation",
                "--output",
                str(receipt),
            ]
            with (
                patch.object(sys, "argv", argv),
                patch.object(verify_workbook, "verify_precalculation", return_value=report),
                contextlib.redirect_stdout(io.StringIO()),
            ):
                verify_main()
            self.assertEqual(json.loads(receipt.read_text(encoding="utf-8")), report)

            argv[argv.index("precalculation")] = "recalculated"
            with (
                patch.object(sys, "argv", argv),
                patch.object(
                    verify_workbook,
                    "verify_recalculated",
                    side_effect=ValueError("recalculation failed"),
                ),
                contextlib.redirect_stderr(io.StringIO()),
            ):
                with self.assertRaises(SystemExit) as context:
                    verify_main()
            self.assertEqual(context.exception.code, 1)
            failure = json.loads(receipt.read_text(encoding="utf-8"))
            self.assertEqual(failure["stage"], "recalculated")
            self.assertIn("recalculation failed", failure["errors"])

    def test_cli_main_paths_without_optional_receipts(self) -> None:
        fill_argv = [
            "fill_excel.py",
            "--company",
            "Example Foods",
            "--inputs",
            "inputs.json",
            "--template",
            "template.xlsx",
            "--output",
            "output.xlsx",
            "--run-id",
            "no-receipt",
        ]
        with (
            patch.object(sys, "argv", fill_argv),
            patch.object(
                fill_excel,
                "fill_valuation_excel",
                return_value={"status": "awaiting_recalculation"},
            ),
            contextlib.redirect_stdout(io.StringIO()),
        ):
            fill_main()
        with (
            patch.object(sys, "argv", fill_argv),
            patch.object(
                fill_excel,
                "fill_valuation_excel",
                side_effect=ValueError("invalid inputs"),
            ),
            contextlib.redirect_stderr(io.StringIO()),
        ):
            with self.assertRaises(SystemExit) as context:
                fill_main()
        self.assertEqual(context.exception.code, 1)

        verify_argv = [
            "verify_workbook.py",
            "--file",
            "model.xlsx",
            "--inputs",
            "inputs.json",
            "--template",
            "template.xlsx",
            "--stage",
            "precalculation",
        ]
        with (
            patch.object(sys, "argv", verify_argv),
            patch.object(
                verify_workbook,
                "verify_precalculation",
                return_value={"stage": "precalculation", "status": "passed"},
            ),
            contextlib.redirect_stdout(io.StringIO()),
        ):
            verify_main()

    def test_fill_module_main_guard_executes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            normalized_path = self._normalize(directory, load_valid_inputs())
            output_path = directory / "guard.xlsx"
            receipt_path = directory / "receipt.json"
            argv = [
                "fill_excel.py",
                "--company",
                "Example Foods",
                "--inputs",
                str(normalized_path),
                "--template",
                str(TEMPLATE),
                "--output",
                str(output_path),
                "--run-id",
                "guard-test",
                "--receipt",
                str(receipt_path),
            ]
            with (
                patch.object(sys, "argv", argv),
                contextlib.redirect_stdout(io.StringIO()),
            ):
                runpy.run_module("fill_excel", run_name="__main__")
            self.assertTrue(output_path.exists())
            self.assertEqual(
                json.loads(receipt_path.read_text(encoding="utf-8"))["run_id"],
                "guard-test",
            )

    def test_verify_module_main_guard_failure(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            invalid_inputs = Path(temporary) / "invalid.json"
            invalid_inputs.write_text("[]", encoding="utf-8")
            argv = [
                "verify_workbook.py",
                "--file",
                str(TEMPLATE),
                "--inputs",
                str(invalid_inputs),
                "--template",
                str(TEMPLATE),
                "--stage",
                "precalculation",
            ]
            with (
                patch.object(sys, "argv", argv),
                contextlib.redirect_stderr(io.StringIO()),
            ):
                with self.assertRaises(SystemExit) as context:
                    runpy.run_module("verify_workbook", run_name="__main__")
            self.assertEqual(context.exception.code, 1)


if __name__ == "__main__":
    unittest.main()
