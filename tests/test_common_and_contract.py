from __future__ import annotations

import contextlib
import io
import json
import math
import runpy
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from common import atomic_write_json, file_sha256, is_finite_number, nested_get, numbers_match
from workbook_contract import (
    ContractValueError,
    _contract_to_json,
    _load_industry_distributions,
    _non_empty_values,
    _validate_template_structure,
    canonicalize_country,
    canonicalize_industry,
    canonicalize_rating,
    canonicalize_region,
    load_workbook_contract,
    protected_formula_fingerprint,
)

from tests.helpers import ROOT, TEMPLATE


class CommonTests(unittest.TestCase):
    def test_common_helpers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "nested" / "value.json"
            atomic_write_json(path, {"b": 2, "a": 1})
            self.assertEqual(json.loads(path.read_text()), {"a": 1, "b": 2})
            self.assertEqual(file_sha256(path), file_sha256(path))
        self.assertEqual(nested_get({"a": {"b": 3}}, "a.b"), 3)
        with self.assertRaises(KeyError):
            nested_get({"a": {}}, "a.missing")
        self.assertTrue(is_finite_number(1.5))
        self.assertFalse(is_finite_number(True))
        self.assertFalse(is_finite_number(math.inf))
        self.assertTrue(numbers_match(1.0, 1.0 + 1e-10))
        self.assertFalse(numbers_match(1.0, 1.1))


class WorkbookContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.contract = load_workbook_contract(TEMPLATE)

    def test_contract_loads_and_serializes(self) -> None:
        self.assertEqual(len(self.contract.industries_global), 94)
        self.assertIn("United States", self.contract.countries)
        self.assertIn("North America", self.contract.regions)
        self.assertIn("Food Processing", self.contract.industry_distributions)
        workbook = openpyxl.load_workbook(TEMPLATE)
        self.assertEqual(
            protected_formula_fingerprint(workbook),
            self.contract.formula_fingerprint,
        )
        workbook.close()
        serialized = _contract_to_json(self.contract)
        self.assertIsInstance(serialized["industries_us"], list)

    def test_canonicalizers(self) -> None:
        self.assertEqual(
            canonicalize_industry("Food Processing", self.contract),
            "Food Processing",
        )
        self.assertEqual(
            canonicalize_industry("FOOD---PROCESSING", self.contract),
            "Food Processing",
        )
        self.assertEqual(
            canonicalize_industry("Food Processing", self.contract, scope="us"),
            "Food Processing",
        )
        self.assertEqual(
            canonicalize_industry("drug", self.contract),
            "Drugs (Pharmaceutical)",
        )
        self.assertEqual(canonicalize_country("usa", self.contract), "United States")
        self.assertEqual(
            canonicalize_region("central/south america", self.contract),
            "Central and South America",
        )
        self.assertEqual(canonicalize_rating("bbb", self.contract), "Baa2/BBB")
        with self.assertRaises(ContractValueError):
            canonicalize_industry("definitely not an industry", self.contract)
        with self.assertRaisesRegex(ContractValueError, "Closest workbook values"):
            canonicalize_industry("Food Process", self.contract)

    def test_template_structure_rejects_label_change(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bad.xlsx"
            workbook = openpyxl.load_workbook(TEMPLATE)
            workbook["Input sheet"]["A4"] = "Wrong"
            workbook.save(path)
            workbook.close()
            with self.assertRaises(ContractValueError):
                load_workbook_contract(path)

    def test_template_structure_rejects_missing_sheet(self) -> None:
        workbook = openpyxl.load_workbook(TEMPLATE)
        del workbook["Diagnostics"]
        with self.assertRaisesRegex(ContractValueError, "missing required sheets"):
            _validate_template_structure(workbook)
        workbook.close()

    def test_contract_parsers_skip_blank_and_invalid_rows(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet["A1"] = None
        worksheet["A2"] = "   "
        worksheet["A3"] = " Value "
        self.assertEqual(_non_empty_values(worksheet, "A1:A3"), ("Value",))

        distribution = workbook.create_sheet("Input Stat Distributioons")
        distribution["A3"] = "Valid"
        for column, value in enumerate((1, 2, 3, 4, 5, 6), start=3):
            distribution.cell(3, column, value)
        distribution["A4"] = "Invalid"
        distribution["C4"] = "not numeric"
        parsed = _load_industry_distributions(workbook)
        self.assertIn("Valid", parsed)
        self.assertNotIn("Invalid", parsed)
        workbook.close()

    def test_contract_loader_skips_missing_region_label(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "missing-region.xlsx"
            workbook = openpyxl.load_workbook(TEMPLATE)
            workbook["Country equity risk premiums"]["A201"] = None
            workbook.save(path)
            workbook.close()
            contract = load_workbook_contract(path)
            self.assertEqual(len(contract.regions), len(self.contract.regions) - 1)

    def test_workbook_contract_module_main_guard_executes(self) -> None:
        argv = ["workbook_contract.py", "--template", str(TEMPLATE)]
        with (
            patch.object(sys, "argv", argv),
            contextlib.redirect_stdout(io.StringIO()) as stdout,
        ):
            runpy.run_module("workbook_contract", run_name="__main__")
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["template_sha256"], self.contract.template_sha256)


class RepositoryLayoutTests(unittest.TestCase):
    def test_v3_layout_replaces_legacy_root_files(self) -> None:
        required_paths = (
            "SKILL.md",
            "README.md",
            "assets/template.xlsx",
            "references/extraction_contract.md",
            "references/google_delivery.md",
            "references/valuation_rules.md",
            "scripts/fill_excel.py",
            "scripts/get_financial_reports.py",
            "scripts/run_valuation.py",
            "scripts/upload_to_sheets.py",
            "scripts/validate_inputs.py",
            "scripts/verify_workbook.py",
            "scripts/workbook_contract.py",
            "pyproject.toml",
            "uv.lock",
        )
        for relative_path in required_paths:
            with self.subTest(required_path=relative_path):
                self.assertTrue((ROOT / relative_path).is_file())

        legacy_root_files = (
            "fill_excel.py",
            "get_financial_reports.py",
            "template.xlsx",
            "upload_to_sheets.py",
            "valuation_theory.txt",
            "workbook_contract.py",
        )
        for relative_path in legacy_root_files:
            with self.subTest(legacy_root_file=relative_path):
                self.assertFalse((ROOT / relative_path).exists())


if __name__ == "__main__":
    unittest.main()
