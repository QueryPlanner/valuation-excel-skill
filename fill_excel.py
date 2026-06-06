import json
import argparse
import openpyxl
from datetime import datetime
import yfinance as yf

def fill_valuation_excel(company_name, inputs_json_path, template_path, output_path, price=None, rf_rate=None, erp=0.055, ticker=None):
    # Fetch market data
    if ticker:
        try:
            stock = yf.Ticker(ticker)
            if price is None:
                price = stock.fast_info.last_price
            if rf_rate is None:
                tnx = yf.Ticker("^TNX")
                rf_rate = tnx.fast_info.last_price / 100
        except Exception as e:
            print(f"Failed to fetch market data via yfinance: {e}")
    
    price = price if price is not None else 100.0
    rf_rate = rf_rate if rf_rate is not None else 0.045
    
    with open(inputs_json_path, "r") as f:
        inputs = json.load(f)

    if isinstance(inputs, list):
        inputs = inputs[0]
    
    fin_data = inputs.get("financial_data", {})
    single_metrics = inputs.get("single_value_metrics", {})
    cost_of_capital = inputs.get("cost_of_capital_inputs", {})
    r_and_d_details = inputs.get("r_and_d_details", {})
    industry_name = r_and_d_details.get("industry_name", "Drugs (Pharmaceutical)")

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Input sheet']

    def write_cell(cell, value):
        ws[cell] = value

    write_cell('B3', datetime.today().strftime('%m/%d/%Y'))
    write_cell('B4', company_name)
    write_cell('B7', "United States") 
    write_cell('B8', industry_name)
    write_cell('B9', industry_name)

    rev = fin_data.get("Revenues", {})
    ebit = fin_data.get("Operating_income_or_EBIT", {})
    interest = fin_data.get("Interest_expense", {})
    equity = fin_data.get("Book_value_of_equity", {})
    debt = fin_data.get("Book_value_of_debt", {})
    cash = fin_data.get("Cash_and_Marketable_Securities", {})
    cross = fin_data.get("Cross_holdings_and_other_non_operating_assets", {})
    minority = fin_data.get("Minority_interests", {})
    
    tax_rate = single_metrics.get("Effective_tax_rate", 0.25)
    
    write_cell('B11', rev.get("Most_Recent_12_months", 0.0))
    write_cell('C11', rev.get("Last_10K_before_LTM", 0.0))
    write_cell('D11', single_metrics.get("Years_since_last_10K", 1.0))      

    write_cell('B12', ebit.get("Most_Recent_12_months", 0.0))  
    write_cell('C12', ebit.get("Last_10K_before_LTM", 0.0))  

    write_cell('B13', interest.get("Most_Recent_12_months", 0.0))   
    write_cell('C13', interest.get("Last_10K_before_LTM", 0.0))   

    write_cell('B14', equity.get("Most_Recent_12_months", 0.0)) 
    write_cell('C14', equity.get("Last_10K_before_LTM", 0.0)) 

    write_cell('B15', debt.get("Most_Recent_12_months", 0.0))  
    write_cell('C15', debt.get("Last_10K_before_LTM", 0.0))  

    write_cell('B16', 'No')     
    write_cell('B17', 'No')     

    write_cell('B18', cash.get("Most_Recent_12_months", 0.0))  
    write_cell('C18', cash.get("Last_10K_before_LTM", 0.0))  

    write_cell('B19', cross.get("Most_Recent_12_months", 0.0))   
    write_cell('C19', cross.get("Last_10K_before_LTM", 0.0))   

    write_cell('B20', minority.get("Most_Recent_12_months", 0.0))   
    write_cell('C20', minority.get("Last_10K_before_LTM", 0.0))      

    shares_out = single_metrics.get("Number_of_shares_outstanding", 0.0)
    write_cell('B21', shares_out)   
    write_cell('B22', price)  
    write_cell('B23', tax_rate)   
    write_cell('B24', 0.25) # Marginal rate default

    # --- INTRINSIC GROWTH CALCULATION ---
    # IC = Equity + Debt - Cash
    ic_ltm = equity.get("Most_Recent_12_months", 0.0) + debt.get("Most_Recent_12_months", 0.0) - cash.get("Most_Recent_12_months", 0.0)
    ic_10k = equity.get("Last_10K_before_LTM", 0.0) + debt.get("Last_10K_before_LTM", 0.0) - cash.get("Last_10K_before_LTM", 0.0)
    
    after_tax_ebit = ebit.get("Most_Recent_12_months", 0.0) * (1 - tax_rate)
    
    if ic_10k > 0:
        roic = after_tax_ebit / ic_10k
    else:
        roic = 0.15 # Fallback
        
    reinvestment = ic_ltm - ic_10k
    
    if after_tax_ebit > 0:
        reinvestment_rate = reinvestment / after_tax_ebit
    else:
        reinvestment_rate = 0.0 # Fallback
        
    fundamental_growth = max(0.01, min(roic * reinvestment_rate, 0.5)) # Cap between 1% and 50%
    current_margin = ebit.get("Most_Recent_12_months", 0.0) / rev.get("Most_Recent_12_months", 1.0) if rev.get("Most_Recent_12_months", 0.0) > 0 else 0.15
    sales_to_capital = rev.get("Most_Recent_12_months", 0.0) / ic_ltm if ic_ltm > 0 else 1.0
    sales_to_capital = max(0.1, min(sales_to_capital, 5.0)) # Sanity bounds

    write_cell('B26', fundamental_growth)   
    write_cell('B27', current_margin)   
    write_cell('B28', fundamental_growth)   
    write_cell('B29', current_margin)   
    write_cell('B30', 5.0)      
    write_cell('B31', sales_to_capital)      
    write_cell('B32', sales_to_capital)      

    write_cell('B34', rf_rate)     
    write_cell('B35', "='Cost of capital worksheet'!B13")
    write_cell('B37', 'No')

    if 'Stories to Numbers' in wb.sheetnames:
        ws_story = wb['Stories to Numbers']
        ws_story['A2'] = "[Insert Your Company Story Title Here]"
        ws_story['A3'] = "GUIDANCE: Read the 'Management Discussion & Analysis' (MD&A) section in the Annual Report. What is the company's core business model? Who are their competitors? Are they a disruptor or an incumbent? Write a narrative here tying their business strategy to their ability to generate cash flows, drive growth, and manage risk."
        ws_story['G9'] = "GUIDANCE: Link to Growth Story. Look at historical revenue trends and management's future guidance. Why will they grow at this rate?"
        ws_story['G10'] = "GUIDANCE: Link to Profitability. Are they cutting costs or scaling? Compare current margins to industry averages to set a realistic target."
        ws_story['G11'] = "GUIDANCE: Check the effective tax rate in the income statement vs the marginal corporate tax rate of their home country."
        ws_story['G12'] = "GUIDANCE: Link to Capital Efficiency. How much reinvestment (CapEx, R&D, Acquisitions) is needed to drive the growth story above?"
        ws_story['G13'] = "GUIDANCE: Link to Competitive Advantage (Moat). A high ROC requires strong barriers to entry. Do they have them?"
        ws_story['G14'] = "GUIDANCE: Link to Risk Profile. Look at their debt load, geographical exposure, and business cyclicality to justify the Cost of Capital."

    # --- COST OF CAPITAL OVERHAUL ---
    if 'Cost of capital worksheet' in wb.sheetnames:
        ws_coc = wb['Cost of capital worksheet']
        def write_coc(cell, value):
            ws_coc[cell] = value

        write_coc('B11', 'Detailed')
        write_coc('B18', shares_out)
        write_coc('B19', price)
        
        # Regional ERP Mapping
        by_region = inputs.get("revenue_splits", {}).get("by_region", {})
        is_regional = any(k in ["Africa", "Asia", "Australia & New Zealand", "Caribbean", "Central and South America", "Eastern Europe", "Middle East", "North America", "Western Europe", "Rest of the World", "EMEA"] for k in by_region.keys())
        
        # Clear existing regions
        for i in range(5, 18):
            write_coc(f'G{i}', None)
            write_coc(f'H{i}', None)
            
        # Clear existing regional mapping
        for i in range(21, 32):
            write_coc(f'H{i}', None)
            
        if is_regional:
            write_coc('B25', 'Operating regions')
            # Mapping strictly to rows 21-31 which are:
            # 21: Africa, 22: Asia, 23: Australia & NZ, 24: Caribbean, 25: Central/South America
            # 26: Eastern Europe, 27: Middle East, 28: North America, 29: Western Europe, 30: EMEA, 31: Rest of World
            region_row_map = {
                "Africa": 21, "Asia": 22, "Australia & New Zealand": 23, "Caribbean": 24,
                "Central and South America": 25, "Eastern Europe": 26, "Middle East": 27,
                "North America": 28, "Western Europe": 29, "EMEA": 30, "Rest of the World": 31
            }
            for region_name, r_rev in by_region.items():
                if region_name in region_row_map and r_rev > 0:
                    write_coc(f'H{region_row_map[region_name]}', r_rev)
        else:
            write_coc('B25', 'Operating countries')
            row = 5
            for country_name, c_rev in by_region.items():
                if c_rev > 0:
                    if row <= 15:
                        write_coc(f'G{row}', country_name)
                        write_coc(f'H{row}', c_rev)
                        row += 1
                    else:
                        # overflow to Rest of the World
                        current = ws_coc['H16'].value or 0
                        write_coc('H16', current + c_rev)
                        
        # Business Beta Mapping
        by_business = inputs.get("revenue_splits", {}).get("by_business", {})
        if len(by_business) > 1:
            write_coc('B21', 'Multibusiness(Global)')
            # Clear existing dummy entries in G52:H63
            for i in range(52, 64):
                write_coc(f'G{i}', None)
                write_coc(f'H{i}', None)
            
            row = 52
            for biz_name, r_rev in by_business.items():
                if row <= 63 and r_rev > 0:
                    write_coc(f'G{row}', biz_name)
                    write_coc(f'H{row}', r_rev)
                    row += 1
        else:
            write_coc('B21', 'Single Business(Global)')

        write_coc('B24', rf_rate)
        # B25 already set above
        write_coc('B26', erp) 
        write_coc('B30', debt.get("Most_Recent_12_months", 0.0))
        write_coc('B31', interest.get("Most_Recent_12_months", 0.0))
        write_coc('B32', cost_of_capital.get("average_maturity_of_debt_years", 5.0))
        write_coc('B33', 'Synthetic rating')
        write_coc('B36', 1) 
        write_coc('B38', tax_rate)

    wb.save(output_path)
    print(f"Successfully populated {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fill Valuation Excel")
    parser.add_argument("--company", type=str, required=True)
    parser.add_argument("--inputs", type=str, required=True)
    parser.add_argument("--template", type=str, default="template.xlsx")
    parser.add_argument("--output", type=str, required=True)
    parser.add_argument("--price", type=float, default=None, help="Current stock price")
    parser.add_argument("--rf_rate", type=float, default=None, help="Risk free rate (e.g. 0.044 for 4.4%)")
    parser.add_argument("--erp", type=float, default=0.055, help="Equity Risk Premium (e.g. 0.055 for 5.5%)")
    parser.add_argument("--ticker", type=str, default=None, help="Yahoo Finance ticker symbol to automatically fetch price and risk-free rate")
    args = parser.parse_args()
    
    fill_valuation_excel(args.company, args.inputs, args.template, args.output, args.price, args.rf_rate, args.erp, args.ticker)
