from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from common import atomic_write_json, file_sha256
from validate_inputs import InputValidationError, normalize_and_validate_inputs
from workbook_contract import load_workbook_contract, protected_formula_fingerprint

INPUT_SHEET = "Input sheet"
STORY_SHEET = "Stories to Numbers"
COST_OF_CAPITAL_SHEET = "Cost of capital worksheet"
R_AND_D_SHEET_NAMES = ("R& D converter", "R&D converter")
LEASE_SHEET = "Operating lease converter"
AUDIT_SHEET = "Sources & Audit"
METADATA_SHEET = "Run Metadata"


def _load_inputs(inputs_json_path: str | Path) -> dict[str, Any]:
    with Path(inputs_json_path).open(encoding="utf-8") as source:
        inputs = json.load(source)
    if not isinstance(inputs, dict):
        raise InputValidationError(["The valuation inputs JSON must contain an object."])
    return inputs


def _write_input_sheet(worksheet, company_name: str, inputs: dict[str, Any]) -> None:
    context = inputs["company_context"]
    financial_data = inputs["financial_data"]
    single_metrics = inputs["single_value_metrics"]
    base_case = inputs["base_case_assumptions"]
    r_and_d = inputs["r_and_d_details"]
    employee_options = inputs["employee_options"]
    lease = inputs["operating_lease_details"]
    advanced = inputs["advanced_assumptions"]
    market = inputs["market_inputs"]

    worksheet["B3"] = context["valuation_date"]
    worksheet["B4"] = company_name
    worksheet["B7"] = context["country_of_incorporation"]
    worksheet["B8"] = context["industry_us"]
    worksheet["B9"] = context["industry_global"]

    metric_rows = {
        "Revenues": 11,
        "Operating_income_or_EBIT": 12,
        "Interest_expense": 13,
        "Book_value_of_equity": 14,
        "Book_value_of_debt": 15,
        "Cash_and_Marketable_Securities": 18,
        "Cross_holdings_and_other_non_operating_assets": 19,
        "Minority_interests": 20,
    }
    for metric_name, row_number in metric_rows.items():
        metric = financial_data[metric_name]
        worksheet[f"B{row_number}"] = metric["Most_Recent_12_months"]
        worksheet[f"C{row_number}"] = metric["Last_10K_before_LTM"]

    worksheet["D11"] = single_metrics["Years_since_last_10K"]
    worksheet["B16"] = "Yes" if r_and_d["current_year_expense"] > 0 else "No"
    worksheet["B17"] = "Yes" if lease.get("capitalize") else "No"
    worksheet["B21"] = single_metrics["Number_of_shares_outstanding"]
    worksheet["B22"] = market["stock_price"]
    worksheet["B23"] = single_metrics["Effective_tax_rate"]
    worksheet["B24"] = single_metrics["Marginal_tax_rate"]
    worksheet["B26"] = base_case["revenue_growth_next_year"]
    worksheet["B27"] = base_case["operating_margin_next_year"]
    worksheet["B28"] = base_case["revenue_cagr_years_2_5"]
    worksheet["B29"] = base_case["target_operating_margin"]
    worksheet["B30"] = base_case["margin_convergence_year"]
    worksheet["B31"] = base_case["sales_to_capital_years_1_5"]
    worksheet["B32"] = base_case["sales_to_capital_years_6_10"]
    worksheet["B34"] = market["riskfree_rate"]
    worksheet["B35"] = "='Cost of capital worksheet'!B13"

    option_count = employee_options["total_options_outstanding"]
    worksheet["B37"] = "Yes" if option_count > 0 else "No"
    for coordinate in ("B38", "B39", "B40", "B41"):
        worksheet[coordinate] = 0
    if option_count > 0:
        worksheet["B38"] = option_count
        worksheet["B39"] = employee_options["weighted_average_exercise_price"]
        worksheet["B40"] = employee_options["average_maturity_years"]
        worksheet["B41"] = employee_options["stock_price_standard_deviation"]

    terminal_wacc = advanced["terminal_cost_of_capital_override"]
    worksheet["B45"] = "Yes" if terminal_wacc is not None else "No"
    worksheet["B46"] = terminal_wacc if terminal_wacc is not None else "=B35"
    stable_roc = advanced["stable_return_on_capital_override"]
    worksheet["B48"] = "Yes" if stable_roc is not None else "No"
    worksheet["B49"] = stable_roc if stable_roc is not None else 0
    failure_probability = advanced["probability_of_failure"]
    worksheet["B51"] = "Yes" if failure_probability > 0 else "No"
    worksheet["B52"] = failure_probability
    worksheet["B53"] = advanced["failure_proceeds_basis"]
    worksheet["B54"] = advanced["failure_proceeds_percent"]
    lag = advanced["reinvestment_lag_years"]
    worksheet["B56"] = "Yes" if lag != 1 else "No"
    worksheet["B57"] = lag
    worksheet["B59"] = "Yes" if advanced["hold_effective_tax_rate"] else "No"
    net_operating_loss = advanced["net_operating_loss"]
    worksheet["B61"] = "Yes" if net_operating_loss > 0 else "No"
    worksheet["B62"] = net_operating_loss
    stable_riskfree = advanced["stable_riskfree_rate_override"]
    worksheet["B64"] = "Yes" if stable_riskfree is not None else "No"
    worksheet["B65"] = stable_riskfree if stable_riskfree is not None else market["riskfree_rate"]
    terminal_growth = advanced["terminal_growth_override"]
    worksheet["B67"] = "Yes" if terminal_growth is not None else "No"
    worksheet["B68"] = terminal_growth if terminal_growth is not None else market["riskfree_rate"]
    trapped_cash = advanced["trapped_cash"]
    worksheet["B70"] = "Yes" if trapped_cash > 0 else "No"
    worksheet["B71"] = trapped_cash
    worksheet["B72"] = advanced["trapped_cash_tax_rate"]


def _write_story_sheet(worksheet, inputs: dict[str, Any]) -> None:
    story = inputs["company_story"]
    worksheet["A2"] = story["title"]
    worksheet["A3"] = story["summary"]
    for row_number, field in enumerate(
        ("growth", "profitability", "tax", "reinvestment", "competitive_advantage", "risk"),
        start=9,
    ):
        worksheet[f"G{row_number}"] = story[field]


def _clear_cost_of_capital_lists(worksheet) -> None:
    for row_number in range(5, 18):
        worksheet[f"G{row_number}"] = None
        worksheet[f"H{row_number}"] = None
    for row_number in range(21, 32):
        worksheet[f"H{row_number}"] = None
    for row_number in range(36, 48):
        worksheet[f"G{row_number}"] = None
        worksheet[f"H{row_number}"] = None
    for row_number in range(52, 64):
        worksheet[f"G{row_number}"] = None
        worksheet[f"H{row_number}"] = None


def _write_geographic_revenue(
    worksheet,
    inputs: dict[str, Any],
    region_rows: dict[str, int],
) -> str:
    splits = inputs["revenue_splits"]
    mode = splits["geography_mode"]
    if mode == "country":
        by_country = splits["by_country"]
        rest_of_world_revenue = by_country.get("Rest of the World", 0)
        named = [(name, value) for name, value in by_country.items() if name != "Rest of the World"]
        for row_number, (country, revenue) in enumerate(named, start=5):
            worksheet[f"G{row_number}"] = country
            worksheet[f"H{row_number}"] = revenue
        if rest_of_world_revenue:
            worksheet["G16"] = "Rest of the World"
            worksheet["H16"] = rest_of_world_revenue
            worksheet["I16"] = inputs["cost_of_capital_inputs"]["rest_of_world_erp"]
        return "Operating countries"
    if mode == "region":
        for region, revenue in splits["by_region"].items():
            worksheet[f"H{region_rows[region]}"] = revenue
        return "Operating regions"
    return "Country of incorporation"


def _write_business_revenue(worksheet, inputs: dict[str, Any]) -> None:
    by_business = inputs["revenue_splits"]["by_business"]
    if len(by_business) == 1:
        worksheet["B21"] = "Single Business(Global)"
        return
    worksheet["B21"] = "Multibusiness(Global)"
    for row_number, (industry, revenue) in enumerate(by_business.items(), start=52):
        worksheet[f"G{row_number}"] = industry
        worksheet[f"H{row_number}"] = revenue


def _write_cost_of_capital_sheet(
    worksheet,
    inputs: dict[str, Any],
    region_rows: dict[str, int],
) -> None:
    financial_data = inputs["financial_data"]
    single_metrics = inputs["single_value_metrics"]
    market = inputs["market_inputs"]
    cost_inputs = inputs["cost_of_capital_inputs"]
    _clear_cost_of_capital_lists(worksheet)

    worksheet["B11"] = "Detailed"
    worksheet["B18"] = single_metrics["Number_of_shares_outstanding"]
    worksheet["B19"] = market["stock_price"]
    worksheet["B24"] = market["riskfree_rate"]
    erp_mode = _write_geographic_revenue(worksheet, inputs, region_rows)
    direct_erp = cost_inputs.get("direct_erp")
    if direct_erp is not None:
        worksheet["B25"] = "Will Input"
        worksheet["B26"] = direct_erp
    else:
        worksheet["B25"] = erp_mode
        worksheet["B26"] = 0
    _write_business_revenue(worksheet, inputs)

    worksheet["B30"] = financial_data["Book_value_of_debt"]["Most_Recent_12_months"]
    worksheet["B31"] = financial_data["Interest_expense"]["Most_Recent_12_months"]
    worksheet["B32"] = cost_inputs["average_maturity_of_debt_years"]
    if cost_inputs["debt_rating"] != "N/A":
        worksheet["B33"] = "Actual rating"
        worksheet["B35"] = cost_inputs["debt_rating"]
        worksheet["B36"] = 0
    else:
        worksheet["B33"] = "Synthetic Rating"
        worksheet["B35"] = None
        worksheet["B36"] = cost_inputs["synthetic_rating_company_type"]
    worksheet["B38"] = single_metrics["Marginal_tax_rate"]


def _write_r_and_d_sheet(workbook, inputs: dict[str, Any]) -> None:
    sheet_name = next((name for name in R_AND_D_SHEET_NAMES if name in workbook.sheetnames), None)
    if sheet_name is None:
        raise ValueError("The workbook is missing the R&D converter.")
    worksheet = workbook[sheet_name]
    r_and_d = inputs["r_and_d_details"]
    for row_number in range(11, 21):
        worksheet[f"B{row_number}"] = None
    current_expense = r_and_d["current_year_expense"]
    if current_expense <= 0:
        worksheet["F6"] = 1
        worksheet["F7"] = 0
        return
    years = r_and_d["amortization_period_years"]
    worksheet["F6"] = years
    worksheet["F7"] = current_expense
    for year_number in range(1, years + 1):
        worksheet[f"B{10 + year_number}"] = r_and_d["historical_expenses"][f"Year_Minus_{year_number}"]


def _write_lease_sheet(workbook, inputs: dict[str, Any]) -> None:
    worksheet = workbook[LEASE_SHEET]
    worksheet["E4"] = 0
    for row_number in range(7, 13):
        worksheet[f"B{row_number}"] = 0
    lease = inputs["operating_lease_details"]
    if not lease.get("capitalize"):
        return
    worksheet["E4"] = lease["current_year_expense"]
    commitments = lease["commitments"]
    for row_number, key in enumerate(
        ("year_1", "year_2", "year_3", "year_4", "year_5", "years_6_and_beyond"),
        start=7,
    ):
        worksheet[f"B{row_number}"] = commitments[key]


def _replace_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name)


def _style_headers(worksheet, row_number: int, end_column: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet.iter_cols(min_row=row_number, max_row=row_number, min_col=1, max_col=end_column):
        header = cell[0]
        header.fill = fill
        header.font = Font(color="FFFFFF", bold=True)
        header.alignment = Alignment(vertical="center")


def _write_audit_sheet(workbook, inputs: dict[str, Any]) -> None:
    worksheet = _replace_sheet(workbook, AUDIT_SHEET)
    headers = (
        "Metric",
        "Value",
        "Units",
        "Source type",
        "Period",
        "Accession",
        "Source URL",
        "SHA-256",
        "Calculation",
        "Rationale",
        "Section",
        "Reported label",
    )
    worksheet.append(headers)
    _style_headers(worksheet, 1, len(headers))
    for record in inputs["source_evidence"]:
        worksheet.append(
            (
                record.get("metric"),
                record.get("value"),
                record.get("units"),
                record.get("source_type"),
                record.get("period"),
                record.get("accession_number"),
                record.get("source_url"),
                record.get("sha256"),
                record.get("calculation"),
                record.get("rationale"),
                record.get("section"),
                record.get("reported_label"),
            )
        )
    widths = (48, 18, 18, 20, 18, 24, 55, 68, 55, 55, 30, 30)
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False


def _write_metadata_sheet(workbook, inputs: dict[str, Any], run_id: str) -> None:
    worksheet = _replace_sheet(workbook, METADATA_SHEET)
    context = inputs["company_context"]
    rows = (
        ("Run ID", run_id),
        ("Status", "Awaiting recalculation"),
        ("Schema version", inputs["schema_version"]),
        ("Company", context.get("company_name", "")),
        ("Ticker", context["ticker"]),
        ("CIK", context.get("cik", "")),
        ("Valuation date", context["valuation_date"]),
        ("Currency", context["currency"]),
        ("Units", context["units"]),
        ("Template SHA-256", inputs["workbook_contract"]["template_sha256"]),
        ("Formula fingerprint", inputs["workbook_contract"]["formula_fingerprint"]),
        ("Generated UTC", datetime.now(timezone.utc).isoformat()),
        ("Normalized input SHA-256", inputs.get("_normalized_input_sha256", "")),
    )
    worksheet.append(("Property", "Value"))
    _style_headers(worksheet, 1, 2)
    for row in rows:
        worksheet.append(row)
    finding_start = len(rows) + 3
    worksheet.cell(finding_start, 1, "Severity")
    worksheet.cell(finding_start, 2, "Code")
    worksheet.cell(finding_start, 3, "Finding")
    _style_headers(worksheet, finding_start, 3)
    for finding in inputs["validation_findings"]:
        worksheet.append((finding["severity"], finding["code"], finding["message"]))
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 72
    worksheet.column_dimensions["C"].width = 100
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.sheet_view.showGridLines = False


def fill_valuation_excel(
    company_name: str,
    inputs_json_path: str | Path,
    template_path: str | Path,
    output_path: str | Path,
    *,
    run_id: str,
    allow_legacy: bool = False,
    replace: bool = False,
) -> dict[str, Any]:
    destination = Path(output_path)
    if destination.exists() and not replace:
        raise FileExistsError(f"Refusing to overwrite existing workbook {destination}.")

    raw_inputs = _load_inputs(inputs_json_path)
    contract = load_workbook_contract(template_path)
    inputs = normalize_and_validate_inputs(raw_inputs, contract, allow_legacy=allow_legacy)
    inputs["_normalized_input_sha256"] = file_sha256(inputs_json_path)

    workbook = openpyxl.load_workbook(template_path)
    original_fingerprint = protected_formula_fingerprint(workbook)
    _write_input_sheet(workbook[INPUT_SHEET], company_name, inputs)
    _write_story_sheet(workbook[STORY_SHEET], inputs)
    _write_cost_of_capital_sheet(workbook[COST_OF_CAPITAL_SHEET], inputs, contract.region_rows)
    _write_r_and_d_sheet(workbook, inputs)
    _write_lease_sheet(workbook, inputs)
    _write_audit_sheet(workbook, inputs)
    _write_metadata_sheet(workbook, inputs, run_id)

    if protected_formula_fingerprint(workbook) != original_fingerprint:
        raise ValueError("Protected workbook formulas changed while populating inputs.")

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.iterate = True
    workbook.calculation.iterateCount = 100
    workbook.calculation.iterateDelta = 0.0001
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()
    return {
        "status": "awaiting_recalculation",
        "workbook_path": str(destination.resolve()),
        "workbook_sha256": file_sha256(destination),
        "run_id": run_id,
        "template_sha256": contract.template_sha256,
        "formula_fingerprint": contract.formula_fingerprint,
        "validation_findings": inputs["validation_findings"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Populate a v2 valuation workbook")
    parser.add_argument("--company", required=True)
    parser.add_argument("--inputs", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--receipt")
    parser.add_argument("--allow-legacy", action="store_true")
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args()
    try:
        receipt = fill_valuation_excel(
            args.company,
            args.inputs,
            args.template,
            args.output,
            run_id=args.run_id,
            allow_legacy=args.allow_legacy,
            replace=args.replace,
        )
    except (InputValidationError, FileExistsError, KeyError, ValueError) as error:
        print(f"Unable to populate valuation workbook:\n{error}", file=sys.stderr)
        raise SystemExit(1) from error
    if args.receipt:
        atomic_write_json(args.receipt, receipt)
    print(json.dumps(receipt, indent=2))


if __name__ == "__main__":  # pragma: no cover - exercised through main()
    main()
