from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

import openpyxl

from common import atomic_write_json, file_sha256, numbers_match
from workbook_contract import load_workbook_contract, protected_formula_fingerprint

FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
ACTIVE_ERROR_SCAN_SHEETS = (
    "Input sheet",
    "Valuation output",
    "Cost of capital worksheet",
    "Synthetic rating",
    "R& D converter",
    "Operating lease converter",
    "Option value",
)


class WorkbookVerificationError(ValueError):
    def __init__(self, errors: list[str], report: dict[str, Any]):
        self.errors = errors
        self.report = report
        super().__init__("\n".join(f"- {error}" for error in errors))


def _load_inputs(path: str | Path) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as source:
        value = json.load(source)
    if not isinstance(value, dict):
        raise ValueError("Normalized inputs must contain an object.")
    return value


def _compare_cell(errors: list[str], worksheet, coordinate: str, expected: Any) -> None:
    actual = worksheet[coordinate].value
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        if not numbers_match(float(actual), float(expected)):
            errors.append(f"{worksheet.title}!{coordinate} expected {expected!r} but found {actual!r}.")
    elif actual != expected:
        errors.append(f"{worksheet.title}!{coordinate} expected {expected!r} but found {actual!r}.")


def _active_sheets(inputs: dict[str, Any]) -> tuple[str, ...]:
    sheets = ["Input sheet", "Valuation output", "Cost of capital worksheet", "Synthetic rating"]
    if inputs["r_and_d_details"]["current_year_expense"] > 0:
        sheets.append("R& D converter")
    if inputs["operating_lease_details"].get("capitalize"):
        sheets.append("Operating lease converter")
    if inputs["employee_options"]["total_options_outstanding"] > 0:
        sheets.append("Option value")
    return tuple(sheets)


def verify_precalculation(
    workbook_path: str | Path,
    normalized_inputs_path: str | Path,
    template_path: str | Path,
) -> dict[str, Any]:
    inputs = _load_inputs(normalized_inputs_path)
    contract = load_workbook_contract(template_path)
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    errors: list[str] = []

    required_sheets = {"Sources & Audit", "Run Metadata"}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        errors.append(f"Workbook is missing audit sheets: {', '.join(sorted(missing))}.")

    if protected_formula_fingerprint(workbook) != contract.formula_fingerprint:
        errors.append("Protected formula fingerprint does not match the template.")

    calculation = workbook.calculation
    if calculation.calcMode != "auto" or not calculation.fullCalcOnLoad or not calculation.forceFullCalc:
        errors.append("Workbook full-recalculation flags are not enabled.")
    if calculation.iterate is not True:
        errors.append("Workbook iterative calculation is not enabled.")

    input_sheet = workbook["Input sheet"]
    context = inputs["company_context"]
    market = inputs["market_inputs"]
    _compare_cell(errors, input_sheet, "B3", context["valuation_date"])
    _compare_cell(errors, input_sheet, "B7", context["country_of_incorporation"])
    _compare_cell(errors, input_sheet, "B8", context["industry_us"])
    _compare_cell(errors, input_sheet, "B9", context["industry_global"])
    _compare_cell(errors, input_sheet, "B22", market["stock_price"])
    _compare_cell(errors, input_sheet, "B34", market["riskfree_rate"])

    formula_ref_errors: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and "#REF!" in str(cell.value):
                    formula_ref_errors.append(f"{worksheet.title}!{cell.coordinate}")
    if formula_ref_errors:
        errors.append("Formula text contains #REF! at " + ", ".join(formula_ref_errors[:20]) + ".")

    report = {
        "stage": "precalculation",
        "status": "passed" if not errors else "failed",
        "errors": errors,
        "workbook_path": str(Path(workbook_path).resolve()),
        "workbook_sha256": file_sha256(workbook_path),
        "formula_fingerprint": protected_formula_fingerprint(workbook),
        "recalculation_required": True,
    }
    workbook.close()
    if errors:
        raise WorkbookVerificationError(errors, report)
    return report


def verify_recalculated(
    workbook_path: str | Path,
    normalized_inputs_path: str | Path,
    template_path: str | Path,
) -> dict[str, Any]:
    preflight = verify_precalculation(workbook_path, normalized_inputs_path, template_path)
    inputs = _load_inputs(normalized_inputs_path)
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    errors: list[str] = []
    active_sheets = _active_sheets(inputs)
    error_cells: list[str] = []
    for sheet_name in active_sheets:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in FORMULA_ERRORS:
                    error_cells.append(f"{sheet_name}!{cell.coordinate}={cell.value}")
    if error_cells:
        errors.append("Active calculation paths contain formula errors: " + ", ".join(error_cells[:30]) + ".")

    key_outputs = {
        "wacc": workbook["Cost of capital worksheet"]["B13"].value,
        "operating_assets": workbook["Valuation output"]["B24"].value,
        "common_equity": workbook["Valuation output"]["B31"].value,
        "value_per_share": workbook["Valuation output"]["B33"].value,
    }
    for name, value in key_outputs.items():
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            errors.append(f"Calculated output {name} is not a finite number.")
    if isinstance(key_outputs["wacc"], (int, float)) and not 0 < key_outputs["wacc"] < 0.5:
        errors.append("Calculated WACC must be greater than 0 and less than 0.5.")

    report = {
        "stage": "recalculated",
        "status": "passed" if not errors else "failed",
        "errors": errors,
        "workbook_path": str(Path(workbook_path).resolve()),
        "workbook_sha256": file_sha256(workbook_path),
        "active_sheets": list(active_sheets),
        "outputs": key_outputs,
        "recalculation_required": False,
        "precalculation": preflight,
    }
    workbook.close()
    if errors:
        raise WorkbookVerificationError(errors, report)
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Verify a populated or recalculated valuation workbook")
    parser.add_argument("--file", required=True)
    parser.add_argument("--inputs", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--stage", choices=("precalculation", "recalculated"), required=True)
    parser.add_argument("--output")
    args = parser.parse_args()
    try:
        if args.stage == "precalculation":
            report = verify_precalculation(args.file, args.inputs, args.template)
        else:
            report = verify_recalculated(args.file, args.inputs, args.template)
    except (ValueError, WorkbookVerificationError) as error:
        report = getattr(error, "report", {"stage": args.stage, "status": "failed", "errors": [str(error)]})
        if args.output:
            atomic_write_json(args.output, report)
        print(json.dumps(report, indent=2), file=sys.stderr)
        raise SystemExit(1) from error
    if args.output:
        atomic_write_json(args.output, report)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":  # pragma: no cover - exercised through main()
    main()
