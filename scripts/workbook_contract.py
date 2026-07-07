from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from difflib import get_close_matches
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from common import file_sha256

INDUSTRY_ALIASES = {
    "computer & peripherals": "Computers/Peripherals",
    "computer software & svcs": "Software (System & Application)",
    "drug": "Drugs (Pharmaceutical)",
    "drug (10)": "Drugs (Pharmaceutical)",
    "drugs": "Drugs (Pharmaceutical)",
    "healthcare info systems": "Heathcare Information and Technology",
    "medical services": "Healthcare Support Services",
    "medical supplies": "Healthcare Products",
    "restaurant": "Restaurant/Dining",
    "semiconductor cap equip": "Semiconductor Equip",
}

COUNTRY_ALIASES = {
    "u.s.": "United States",
    "u.s.a.": "United States",
    "uk": "United Kingdom",
    "united states of america": "United States",
    "us": "United States",
    "usa": "United States",
}

REGION_ALIASES = {
    "australia and new zealand": "Australia & New Zealand",
    "central & south america": "Central and South America",
    "central/south america": "Central and South America",
    "rest of world": "Rest of the World",
}

RATING_ALIASES = {
    "aaa": "Aaa/AAA",
    "aa+": "Aa2/AA",
    "aa": "Aa2/AA",
    "aa-": "Aa2/AA",
    "a+": "A1/A+",
    "a": "A2/A",
    "a-": "A3/A-",
    "bbb+": "Baa2/BBB",
    "bbb": "Baa2/BBB",
    "bbb-": "Baa2/BBB",
    "bb+": "Ba1/BB+",
    "bb": "Ba2/BB",
    "b+": "B1/B+",
    "b": "B2/B",
    "b-": "B3/B-",
    "ccc+": "Caa/CCC",
    "ccc": "Caa/CCC",
    "ccc-": "Caa/CCC",
    "cc": "Ca2/CC",
    "c": "C2/C",
    "d": "D2/D",
}

REQUIRED_SHEETS = (
    "Input sheet",
    "Valuation output",
    "Stories to Numbers",
    "Diagnostics",
    "Option value",
    "Synthetic rating",
    "R& D converter",
    "Operating lease converter",
    "Cost of capital worksheet",
    "Failure Rate worksheet",
    "Country equity risk premiums",
    "Industry Averages(US)",
    "Industry Averages (Global)",
    "Input Stat Distributioons",
    "Answer keys",
)

EXPECTED_LABELS = {
    ("Input sheet", "A4"): "Company name",
    ("Input sheet", "A11"): "Revenues",
    ("Valuation output", "A33"): "Estimated value /share",
    ("Cost of capital worksheet", "A13"): "Cost of capital based upon approach =",
    ("R& D converter", "A6"): "Over how many years do you want to amortize R&D expenses",
    ("Operating lease converter", "A4"): "Operating lease expense in current year =",
}

PROTECTED_FORMULA_SHEETS = (
    "Valuation output",
    "Diagnostics",
    "Option value",
    "Synthetic rating",
    "Failure Rate worksheet",
    "Country equity risk premiums",
)


class ContractValueError(ValueError):
    pass


@dataclass(frozen=True)
class IndustryDistribution:
    growth_q1: float
    growth_median: float
    growth_q3: float
    margin_q1: float
    margin_median: float
    margin_q3: float


@dataclass(frozen=True)
class WorkbookContract:
    template_sha256: str
    formula_fingerprint: str
    industries_us: tuple[str, ...]
    industries_global: tuple[str, ...]
    countries: tuple[str, ...]
    regions: tuple[str, ...]
    ratings: tuple[str, ...]
    region_rows: dict[str, int]
    industry_distributions: dict[str, IndustryDistribution]


def _non_empty_values(worksheet, cell_range: str) -> tuple[str, ...]:
    values: list[str] = []
    for row in worksheet[cell_range]:
        value = row[0].value
        if value is None:
            continue
        text = str(value).strip()
        if text:
            values.append(text)
    return tuple(values)


def protected_formula_fingerprint(workbook) -> str:
    formulas: list[str] = []
    for sheet_name in PROTECTED_FORMULA_SHEETS:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas.append(f"{sheet_name}!{cell.coordinate}={cell.value}")
    payload = "\n".join(sorted(formulas)).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _validate_template_structure(workbook) -> None:
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ContractValueError(f"Template is missing required sheets: {', '.join(missing_sheets)}.")

    label_errors: list[str] = []
    for (sheet_name, coordinate), expected in EXPECTED_LABELS.items():
        actual = workbook[sheet_name][coordinate].value
        if actual != expected:
            label_errors.append(
                f"{sheet_name}!{coordinate} expected {expected!r} but found {actual!r}"
            )
    if label_errors:
        raise ContractValueError("Template labels do not match the v2 contract: " + "; ".join(label_errors))


def _load_industry_distributions(workbook) -> dict[str, IndustryDistribution]:
    worksheet = workbook["Input Stat Distributioons"]
    distributions: dict[str, IndustryDistribution] = {}
    for row_number in range(3, 98):
        name = worksheet[f"A{row_number}"].value
        values = [worksheet.cell(row_number, column).value for column in range(3, 9)]
        if not name or any(not isinstance(value, (int, float)) for value in values):
            continue
        distributions[str(name).strip()] = IndustryDistribution(*(float(value) for value in values))
    return distributions


def load_workbook_contract(template_path: str | Path) -> WorkbookContract:
    workbook = load_workbook(template_path, read_only=False, data_only=False)
    _validate_template_structure(workbook)

    industries_us = _non_empty_values(workbook["Industry Averages(US)"], "A2:A95")
    industries_global = _non_empty_values(workbook["Industry Averages (Global)"], "A2:A95")
    countries = _non_empty_values(workbook["Country equity risk premiums"], "A5:A196")
    ratings = _non_empty_values(workbook["Answer keys"], "G2:G16")

    country_risk = workbook["Country equity risk premiums"]
    region_rows: dict[str, int] = {}
    for row_number in range(21, 32):
        region_name = country_risk[f"A{180 + row_number}"].value
        if region_name:
            region_rows[str(region_name).strip()] = row_number

    contract = WorkbookContract(
        template_sha256=file_sha256(template_path),
        formula_fingerprint=protected_formula_fingerprint(workbook),
        industries_us=industries_us,
        industries_global=industries_global,
        countries=countries,
        regions=tuple(region_rows),
        ratings=ratings,
        region_rows=region_rows,
        industry_distributions=_load_industry_distributions(workbook),
    )
    workbook.close()
    return contract


def _normalized_label(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()


def _canonicalize(
    value: str,
    valid_values: Iterable[str],
    aliases: dict[str, str],
    field_name: str,
) -> str:
    text = str(value).strip()
    valid = tuple(valid_values)
    if text in valid:
        return text

    normalized_to_valid = {_normalized_label(item): item for item in valid}
    normalized_text = _normalized_label(text)
    if normalized_text in normalized_to_valid:
        return normalized_to_valid[normalized_text]

    alias_target = aliases.get(text.casefold()) or aliases.get(normalized_text)
    if alias_target in valid:
        return alias_target

    suggestions = get_close_matches(text, valid, n=3, cutoff=0.45)
    suggestion_text = f" Closest workbook values: {', '.join(suggestions)}." if suggestions else ""
    raise ContractValueError(
        f"{field_name} value {text!r} is not present in the workbook contract.{suggestion_text}"
    )


def canonicalize_industry(value: str, contract: WorkbookContract, scope: str = "global") -> str:
    valid_values = contract.industries_global if scope == "global" else contract.industries_us
    return _canonicalize(value, valid_values, INDUSTRY_ALIASES, f"{scope} industry")


def canonicalize_country(value: str, contract: WorkbookContract) -> str:
    return _canonicalize(value, contract.countries, COUNTRY_ALIASES, "country")


def canonicalize_region(value: str, contract: WorkbookContract) -> str:
    return _canonicalize(value, contract.regions, REGION_ALIASES, "region")


def canonicalize_rating(value: str, contract: WorkbookContract) -> str:
    return _canonicalize(value, contract.ratings, RATING_ALIASES, "debt rating")


def _contract_to_json(contract: WorkbookContract) -> dict[str, object]:
    value = asdict(contract)
    value["industries_us"] = list(contract.industries_us)
    value["industries_global"] = list(contract.industries_global)
    value["countries"] = list(contract.countries)
    value["regions"] = list(contract.regions)
    value["ratings"] = list(contract.ratings)
    return value


def main() -> None:
    parser = argparse.ArgumentParser(description="Print and validate the v2 workbook contract")
    parser.add_argument("--template", required=True, help="Path to the valuation workbook")
    args = parser.parse_args()
    print(json.dumps(_contract_to_json(load_workbook_contract(args.template)), indent=2))


if __name__ == "__main__":  # pragma: no cover - exercised through main()
    main()
